#!/usr/bin/python
import openpyxl
import os
import time
import numpy as np
import matplotlib.pyplot as plt

#Linear regression algorithm
def estimate_coef(x, y):
    # number of observations/points
    n = np.size(x)

    # mean of x and y vector
    m_x, m_y = np.mean(x), np.mean(y)

    # calculating cross-deviation and deviation about x
    SS_xy = np.sum(y*x) - n*m_y*m_x
    SS_xx = np.sum(x*x) - n*m_x*m_x

    # calculating regression coefficients
    b_1 = SS_xy / SS_xx
    b_0 = m_y - b_1*m_x

    return(b_0, b_1)

def plot_regression_line(x, y, b):
    # plotting the actual points as scatter plot
    plt.scatter(x, y, color = "m",
               marker = "o", s = 30)

    # predicted response vector
    y_pred = b[0] + b[1]*x

    # plotting the regression line
    plt.plot(x, y_pred, color = "g")

    # putting labels
    plt.xlabel('x')
    plt.ylabel('y')

    # function to show plot
    plt.show()

def main():
    begin = time.ctime()
    current = os.getcwd()

    file1 = os.path.join(current, 'DATA', 'acceleration_data_1_CRACK_0.0mm.xlsx')
#    file2 = os.path.join(current, 'DATA', 'acceleration_data_1_CRACK_1.5mm.xlsx')
#    file3 = os.path.join(current, 'DATA', 'acceleration_data_1_CRACK_3.0mm.xlsx')
#    file4 = os.path.join(current, 'DATA', 'acceleration_data_1_CRACK_4.5mm.xlsx')
    #sample=openpyxl.load_workbook('acceleration_data_1_CRACK_0mm.xlsx')
    fileOne = openpyxl.load_workbook(file1)
#    fileTwo = openpyxl.load_workbook(file2)
#    fileThree = openpyxl.load_workbook(file3)
#    fileFour = openpyxl.load_workbook(file4)
    sheet1 = fileOne.active
#    sheet2 = fileTwo.active
#    sheet3 = fileThree.active
#    sheet4 = fileFour.active
    m_row = sheet1.max_row

    timelist = []
    blist = []
    clist = []
    #populates the time array
    for i in range(2,m_row + 1):
        cell = sheet1.cell(row = i, column = 2)
        timelist.append(cell.value)
    #populates an array for the first variable in the first document
    for i in range(2,m_row + 1):
        cell = sheet1.cell(row = i, column = 3)
        blist.append(cell.value)

    for i in range(2,m_row + 1):
        cell = sheet1.cell(row = i, column = 10)
        clist.append(cell.value)


    end = time.ctime()
    #print(alist)
    print(blist)
    print(timelist)
    print(begin)
    print(end)
    # observations
    x = np.array(blist)
    y = np.array(clist)

    # estimating coefficients
    b = estimate_coef(x, y)
    print("Estimated coefficients:\nb_0 = {}  \
          \nb_1 = {}".format(b[0], b[1]))

    # plotting regression line
    plot_regression_line(x, y, b)

if __name__ == "__main__":
    main()
