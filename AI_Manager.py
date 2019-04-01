# AI_Manager.py

import sys
import openpyxl
from typing import List
from math import sqrt
from statistics import median
from scipy import fftpack
from scipy import signal
import threading
from Decision_Tree import Decision_Tree

"""
This function loads the xl file given into a 2D list
"""
def process_workbook( input_workbook: openpyxl.Workbook ):

	sheet = input_workbook.active

	intermediate_array = [ column for column in sheet.iter_rows( min_row = 1, max_col = sheet.max_column, max_row = sheet.max_row, values_only = True ) ]

	return_data = []

	for column in range( sheet.max_column ):
		line = [ row[ column ] for row in intermediate_array if isinstance( row[ column ], float ) ]

		if line:
			return_data.append( line )

	return return_data

"""
This class loads and runs each AI.
"""
class AI_Manager:

	def __init__( self, load = True ):
		self.__feature_list__ = None
		self.__restraint__ = 20
		self.__AIs__ = [ Decision_Tree( load_tree = load ) ]

	def set_restraint( self, input_value: int ):
		if 0 < input_value <= 20:
			self.__restraint__ = input_value

	def set_features( self, input_list: List[ float ], time_start: float, time_end: float ):
		self.__feature_list__ = self.__Statistical_Features__( input_list, time_start, time_end )

	def get_feature_list( self ):
		return self.__feature_list__[ : self.__restraint__ ]

	def is_loaded( self ):
		return [ ai.__class__.__name__ for ai in self.__AIs__ if ai.is_loaded() ]

	def is_not_loaded( self ):
		return [ ai.__class__.__name__ for ai in self.__AIs__ if not ai.is_loaded() ]

	"""
	This function automates the getting a result from each AI from a set of lists.
	A set of results is returned to be analyzed
	"""
	def Test_AIs( self, input_list: List[ List[ List[ float ] ] ], expected_result_list: List[ List[ float ] ] ):

		Precision_Stats = {}
		unique_expected_results = set( [ item for sublist in expected_result_list for item in sublist ] )

		for ai in self.__AIs__:
			Precision_Stats.update( { ai.__class__.__name__: { value : {} for value in unique_expected_results } } )

		for _2D_gear_list, _2D_expected_result in zip( input_list, expected_result_list ):
			test_time_start = min( _2D_gear_list[ 0 ] )
			test_time_end = max( _2D_gear_list[ 0 ] )

			for gear_list, expected_result in zip( _2D_gear_list[ 1 : ], _2D_expected_result ):
				self.set_features( gear_list, test_time_start, test_time_end )

				for ai, test_result in self.GetAllResults().items():
					if test_result in Precision_Stats[ ai ][ expected_result ].keys():
						Precision_Stats[ ai ][ expected_result ][ test_result ] += 1
					else:
						Precision_Stats[ ai ][ expected_result ][ test_result ] = 1

		return Precision_Stats

	"""
	This function automates and runs concurrently the training function for each AI
	"""
	def Train_AIs( self, input_list: List[ List[ List[ float ] ] ], expected_result_list: List[ List[ float ] ]  ):

		test_feature_list = []

		for _2D_gear_list in input_list:
			test_time_start = min( _2D_gear_list[ 0 ] )
			test_time_end = max( _2D_gear_list[ 0 ] )

			test_feature_list += [ self.__Statistical_Features__( gear_list, test_time_start, test_time_end )[ : self.__restraint__ ] for gear_list in _2D_gear_list[ 1 : ] ]

		expected_results = [ item for sublist in expected_result_list for item in sublist ]

		ai_threads = [ threading.Thread( target = ai.train, args = ( test_feature_list, expected_results ) ) for ai in self.__AIs__ ]

		for ai_thread in ai_threads:
			ai_thread.start()

		for ai_thread in ai_threads:
			ai_thread.join()

	def Determine_Best_Result( self ):
		# determine AI to use

		# get result from AI (first one for now)
		return self.__AIs__[ 0 ].classify( self.get_feature_list() )

	"""
	This function uses the classify function in each AI to get and return a result for each.
	It requires that each AI to be loaded and the feature list to be set.
	"""
	def GetAllResults( self ):
		results = {}
		feature_list = self.get_feature_list()

		if feature_list:
			ai_threads = [ threading.Thread( target = lambda: \
				results.update( { ai.__class__.__name__: ai.classify( feature_list ) } ) ) \
					for ai in self.__AIs__ if ai.is_loaded() ]

			for ai_thread in ai_threads:
				ai_thread.start()

			for ai_thread in ai_threads:
				ai_thread.join()

		return results

	"""
	This function gets the statistical features of the acceleration data given to it.
	The output is a list of 20 length.
	"""
	@staticmethod
	def __Statistical_Features__( input_list: List[ float ], time_start: float, time_end: float ):

		"""
		div-zero liabilities
			Length
			1 / value for any value in input_list
			Sums[ 0 ]
			Sums[ 1 ]
			Sums[ 2 ]
			Sums[ 5 ]
			Frequency_Sums[ 0 ]
			sum( [ P1[ index ] for index in signal.find_peaks( P1 )[ 0 ] ] )
		"""

		Length = len( input_list )
		duration = time_end - time_start

		# initial error catch
		if duration == 0 or Length == 0:
			return [ 0 ] * 20

		frequency_length = ( 1 + Length ) // 2

		sampling_frequency = Length / ( time_end - time_start )
		P1 = [ abs( value / Length ) for value in fftpack.fft( input_list )[ : frequency_length ] ]

		for index in range( 1, len( P1 ) - 1 ):
			P1[ index ] *= 2

		frequencies_list = [ sampling_frequency * value / Length for value in range( frequency_length ) ]


		input_list = sorted( input_list )

		Median = median( input_list )

		# F2
		def mean( input_list_ ):
			length_ = len( input_list_ )

			if length_ == 0:
				return 0

			return sum( input_list_ ) / length_

		Mean = mean( input_list )

		Sums = [ 0 ] * 8

		Harmonic_Zero = False

		# combined sum loops									Used in:
		for value in input_list:
			temp_difference = value - Mean

			Sums[ 0 ] += value * value						# Root_Mean_Square, ( Crest_Factor, Peak2RMS, Shape_Factor )

			if value == 0:
				Harmonic_Zero = True
			else:
				Sums[ 1 ] += 1 / value						# Harmonic_Mean

			Sums[ 2 ] += abs( value )						# Shape_Factor
			Sums[ 3 ] += abs( temp_difference )				# Mean_Absolute_Deviation
			Sums[ 4 ] += abs( value - Median )				# Median_Absolute_Deviation

			temp_value = temp_difference * temp_difference

			Sums[ 5 ] += temp_value							# Variance, Skewness, ( Kurtosis, Standard_Deviation )

			temp_value *= temp_difference

			Sums[ 6 ] += temp_value							# Skewness
			Sums[ 7 ] += temp_value * temp_difference		# Kurtosis


		# F0
		Maximum_Value = input_list[ -1 ]

		# F1
		Minimum_Value = input_list[ 0 ]


		# F3
			# F0 - F1
		Peak_to_Peak = Maximum_Value - Minimum_Value

		# F4
		if Harmonic_Zero:
			Harmonic_Mean = 0
		elif Sums[ 1 ] == 0:
			Harmonic_Mean = float( 'inf' )
		else:
			Harmonic_Mean = Length / Sums[ 1 ]

		# F5
			# Mean excluding outliers (10% trimmed mean)

			# If a data point is within OUTLIER_CRITERIA * IQR of the median
			# then it is not an outlier
		OUTLIER_CRITERIA = 1.5 # lower excludes more

		first_quartile = median( input_list[ :Length // 2 ] )
		last_quartile = median( input_list[ Length // 2 + 1: ] )
		inter_quartile_range = last_quartile - first_quartile

		Trimmed_Mean = mean( [ value for value in input_list \
			if not abs( value - Median ) > OUTLIER_CRITERIA * inter_quartile_range ] )

		# F6
		Variance = Sums[ 5 ] / Length

		# F7
		Standard_Deviation = sqrt( Variance )

		# F8
		Mean_Absolute_Deviation = Sums[ 3 ] / Length

		# F9
		Median_Absolute_Deviation = Sums[ 4 ] / Length


		# F15
		Root_Mean_Square = sqrt( Sums[ 0 ] / Length )


		# F10
		if Root_Mean_Square == 0:
			Crest_Factor = float( 'inf' )
		else:
			Crest_Factor = Maximum_Value / Root_Mean_Square

		# F11
			# Peak_to_Peak_Root_Mean_Square
		if Root_Mean_Square == 0:
			Peak2RMS = float( 'inf' )
		else:
			Peak2RMS = max( abs( Minimum_Value ), abs( Maximum_Value ) ) / Root_Mean_Square

		# F12
		if Sums[ 5 ] == 0:
			Skewness = 0
		else:
			Skewness = Sums[ 6 ] / Sums[ 5 ]

		# F13
		if Sums[ 5 ] == 0:
			Kurtosis = float( 'inf' )
		else:
			Kurtosis = Sums[ 7 ] / ( Length * Variance ** 2 )

		# F14
		if Sums[ 2 ] == 0:
			Shape_Factor = float( 'inf' )
		else:
			Shape_Factor = Root_Mean_Square * Length / Sums[ 2 ]


		Frequency_Sums = [ 0 ] * 3

		for value, frequency in zip( P1, frequencies_list ):
			temp_value = value * frequency
			Frequency_Sums[ 0 ] += value
			Frequency_Sums[ 1 ] += temp_value
			Frequency_Sums[ 2 ] += temp_value * frequency

		# F17
		Mean_Frequency = Frequency_Sums[ 0 ] / frequency_length

		# F18
		if Frequency_Sums[ 0 ] == 0:
			Frequency_Center = float( 'inf' )
		else:
			Frequency_Center = Frequency_Sums[ 1 ] / Frequency_Sums[ 0 ]

		# F19
		if Frequency_Sums[ 0 ] == 0:
			Root_Mean_Square_Frequency = float( 'inf' )
		else:
			Root_Mean_Square_Frequency = sqrt( Frequency_Sums[ 2 ] / Frequency_Sums[ 0 ] )

		# F20
		peaks_sum = sum( [ P1[ index ] for index in signal.find_peaks( P1 )[ 0 ] ] )

		if peaks_sum == 0:
			Figure_of_Merit = float( 'inf' )
		else:
			Figure_of_Merit = ( Maximum_Value - Mean ) / peaks_sum

		return [ Crest_Factor,				# 10
				Root_Mean_Square_Frequency,	# 18
				Peak2RMS,					# 11
				Skewness,					# 12
				Mean_Frequency,				# 16
				Minimum_Value,				#  1
				Kurtosis,					# 13
				Frequency_Center,			# 17
				Median_Absolute_Deviation,	#  9
				Mean,						#  2
				Shape_Factor,				# 14
				Peak_to_Peak,				#  3
				Trimmed_Mean,				#  5
				Root_Mean_Square,			# 15
				Harmonic_Mean,				#  4
				Variance,					#  6
				Maximum_Value,				#  0
				Mean_Absolute_Deviation,	#  8
				Standard_Deviation,			#  7
				Figure_of_Merit ]			# 19

"""
For future use
"""
class Result_Stats:
	def __init__( self ):
		self.true_positive  = 0
		self.true_negative  = 0
		self.false_positive = 0
		self.false_negative = 0

	def Accuracy( self ):
		return ( self.true_positive + self.true_negative ) \
			/ (self. true_positive + self.true_negative + self.false_positive + self.false_negative )

	def Precision( self ):
		return self.true_positive/( self.true_positive + self.false_positive )

	def Recall( self ):
		return self.true_positive/( self.true_positive + self.false_negative )

	def F1_Score( self ):
		return ( 2 * self.Recall() * self.Precision() )\
			   /( self.Recall() + self.Precision() )

"""
The primary automator of training and testing the AIs.
It prompts the user for action,loads the files given as arguments
	and runs the AI support functions
"""
def main( arguments: List ):
	prompt = input( "Menu:\n\t1)\tTraining\n\t2)\tTesting\n\tOther)\tQuit\n" )

	if prompt == '1' or prompt == '2':

		verbose = False

		if "-v" in arguments:
			arguments.remove( "-v" )
			verbose = True
		elif "-verbose" in arguments:
			arguments.remove( "-verbose" )
			verbose = True

		if not arguments:
			print( "No file given" )
			return

		data_array = []
		results_array = []

		for arg in arguments:
			if verbose: print( "Loading from", arg )

			try:
				# Open file and load workbook
				workbook = openpyxl.load_workbook( arg, read_only = True )
				workbook_data = process_workbook( workbook )

				wb_length = len( workbook_data )

				if wb_length > 1:
					expected_crack_size = 0

					try:
						expected_crack_size = float( arg.split( 'CRACK_' )[ 1 ].split( 'mm' )[ 0 ] )
					except Exception:
						pass

					data_array.append( workbook_data )
					results_array.append( [ expected_crack_size ] * ( wb_length - 1 ) )

			except openpyxl.utils.exceptions.InvalidFileException:
				if verbose: print( "Error: cannot load from", arg )

		if not data_array:
			print( "Files not usable" )
			return

		if verbose:
			print( "Loading AIs" )

		manager = AI_Manager()

		if not manager.__AIs__:
			print( "Error: No AIs loaded" )
			return

		if prompt == '1':
			if verbose: print( "Starting Training" )

			manager.Train_AIs( data_array, results_array )

			if verbose: print( "Finished" )
		else:
			loaded = True

			for ai in manager.__AIs__:
				if not ai.is_loaded():
					loaded = False
					print( "Error: {} is not loaded".format( ai.__class__.__name__ ) )

			if not loaded:
				return

			if verbose: print( "Starting Testing" )

			confusion_matrix_set = manager.Test_AIs( data_array, results_array )

			# print the confusion matrix
			for ai in manager.__AIs__:
				confusion_matrix = confusion_matrix_set[ ai.__class__.__name__ ]
				expected_lables = sorted( confusion_matrix )

				# extract a set of keys
				classified_lables = sorted( set( [ item for sublist in [ list( confusion_matrix[ key ].keys() ) for key in confusion_matrix ] for item in sublist ] ) )

				# keys each as a string
				classified_lable_strings = [ ", ".join( str( item ) for item in sublist ) for sublist in classified_lables ]

				# maximum lengths for table setup
				max_left_string = max( max( len( str( lable ) ) for lable in expected_lables ), 11 )
				max_string = max( len( lable ) for lable in classified_lable_strings )
				separator = ( max_string + 3 ) * len( classified_lables )

				printing_string = "\n{:" + str( max_left_string + 1 ) + "}\n{}{:>" + str( separator - 1 ) + "}\n{:" + str( max_left_string ) + "}{}"
				printing_lables_string = " | {:" + str( max_string ) + "}"

				# print top lables
				print( printing_string.format( ai.__class__.__name__.title(), "Actual Label", "Classified Label", "" , "".join( [ printing_lables_string.format( lable ) for lable in classified_lable_strings ] ) ) )

				printing_string = "{}\n{:" + str( max_left_string ) + "}{}"
				printing_results_string = " | {:." + str( max_string - 2 ) + "f}"
				separator += max_left_string

				for expected_result in expected_lables:
					given_sum = sum( confusion_matrix[ expected_result ].values() )

					if not given_sum:
						given_sum = 1

					# a set of each result
					result_set = [ confusion_matrix[ expected_result ][ given_result ] / given_sum if given_result in confusion_matrix[ expected_result ].keys() else 0 for given_result in classified_lables ]

					print( printing_string.format( '-' * separator, expected_result, "".join( printing_results_string.format( printed_result ) for printed_result in result_set ) ) )


if __name__ == "__main__":
	main( sys.argv[ 1: ] )