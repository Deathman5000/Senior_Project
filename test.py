#!/usr/bin/python
import openpyxl
import os
import time

begin = time.ctime()
current = os.getcwd()

file1 = os.path.join(current, 'DATA', 'acceleration_data_1_CRACK_0.0mm.xlsx')
file2 = os.path.join(current, 'DATA', 'acceleration_data_1_CRACK_1.5mm.xlsx')
file3 = os.path.join(current, 'DATA', 'acceleration_data_1_CRACK_3.0mm.xlsx')
file4 = os.path.join(current, 'DATA', 'acceleration_data_1_CRACK_4.5mm.xlsx')
#sample=openpyxl.load_workbook('acceleration_data_1_CRACK_0mm.xlsx')
fileOne = openpyxl.load_workbook(file1)
fileTwo = openpyxl.load_workbook(file2)
fileThree = openpyxl.load_workbook(file3)
fileFour = openpyxl.load_workbook(file4)
sheet1 = fileOne.active
sheet2 = fileTwo.active
sheet3 = fileThree.active
sheet4 = fileFour.active
m_row = sheet1.max_row

timelist = []
blist = []
#populates the time array
for i in range(2,m_row + 1):
    cell = sheet1.cell(row = i, column = 2)
    timelist.append(cell.value)
#populates an array for the first variable in the first document
for i in range(2,m_row + 1):
    cell = sheet1.cell(row = i, column = 3)
    blist.append(cell.value)


end = time.ctime()
#print(alist)
print(blist)
print(timelist)
print(begin)
print(end)
